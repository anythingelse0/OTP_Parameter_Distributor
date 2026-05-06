#!/usr/bin/env python3
"""
Dynamic Signal Parser & Parameter Distributor
从输入的信号列表动态解析，自动生成分段方案和组合逻辑分发器
"""

import re
import os
import time
from datetime import datetime
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional
from enum import Enum

# 保留信号颜色常量
RESERVED_COLOR = "D3D3D3"  # 浅灰色


class SignalType(Enum):
    """信号类型"""
    LOGIC = "logic"
    BIT = "bit"
    WIRE = "wire"
    REG = "reg"


@dataclass
class Signal:
    """信号定义"""
    name: str
    width: int
    signal_type: SignalType = SignalType.LOGIC
    value: Optional[int] = 0
    description: str = ""
    explicit_width: bool = False  # 是否显式声明位宽（如 [1-1:0] 或 [0:0]）
    is_reserved: bool = False  # 是否为保留信号（reserved）

    @property
    def bit_range(self) -> str:
        """返回信号的位宽范围字符串
        
        - 多位信号: [N:0]
        - 显式 1-bit: [0:0]
        - 隐式 1-bit: ""（空字符串）
        """
        if self.width == 1:
            return "[0:0]" if self.explicit_width else ""
        return f"[{self.width-1}:0]"

    @property
    def type_str(self) -> str:
        return self.signal_type.value


class DynamicSignalParser:
    """动态信号解析器"""
    
    # 匹配 Verilog/SystemVerilog 信号声明
    SIGNAL_PATTERN = re.compile(
        r"(?:input|output|inout|wire|reg|logic)?\s+"
        r"(?:logic|bit|wire|reg)?\s*"
        r"(?:\[(\d+):(\d+)\])?\s*"  # 位宽 [high:low]
        r"(\w+)"  # 信号名
        r"(?:\s*[;,])?",  # 可选的分号或逗号
        re.IGNORECASE
    )
    
    def __init__(self):
        self.signals: List[Signal] = []
        self.total_width: int = 0
        self._signal_name_counter: Dict[str, int] = {}  # 跟踪信号名出现次数，用于重复检测

    @staticmethod
    def _is_reserved(name: str) -> bool:
        """检查信号名是否为保留信号（reserved）"""
        return name.lower() == "reserved"

    def _get_unique_signal_name(self, name: str) -> str:
        """
        生成唯一的信号名。
        如果信号名已存在，添加后缀 _0, _1, _2...
        保持原始名首次出现的无需后缀，从第二次开始出现才加后缀。
        """
        if name not in self._signal_name_counter:
            # 首次出现，直接使用原名称
            self._signal_name_counter[name] = 0
            return name
        else:
            # 重复出现，增加计数并添加后缀
            self._signal_name_counter[name] += 1
            suffix = self._signal_name_counter[name]
            return f"{name}_{suffix}"

    def _validate_signal_values(self) -> None:
        """验证信号值的合法性"""
        warnings = []
        for signal in self.signals:
            max_value = (1 << signal.width) - 1
            
            # 检查 1-bit 信号 value > 1
            if signal.width == 1 and signal.value > 1:
                warnings.append(
                    f"  Warning: Signal '{signal.name}' is 1-bit but efuse_default_value=0x{signal.value:X} ({signal.value}). "
                    f"Only bit[0]={signal.value & 1} will be used."
                )
            # 检查多比特信号 value 超出位宽范围
            elif signal.width > 1 and signal.value > max_value:
                warnings.append(
                    f"  Warning: Signal '{signal.name}' is {signal.width}-bit (max 0x{max_value:X}) but "
                    f"efuse_default_value=0x{signal.value:X} ({signal.value}). "
                    f"High bits will be ignored."
                )
        
        if warnings:
            # ANSI color codes for Windows terminal
            RED = '\033[91m'
            YELLOW = '\033[93m'
            BOLD = '\033[1m'
            BG_YELLOW = '\033[43m'
            BG_RED = '\033[41m'
            RESET = '\033[0m'
            
            print(f"\n{BG_YELLOW}{BOLD}{RED}  ⚠️  VALUE VALIDATION WARNINGS  ⚠️  {RESET}")
            print(f"{YELLOW}{'='*80}{RESET}")
            for warning in warnings:
                # Extract signal name for highlighting
                if "Signal '" in warning:
                    parts = warning.split("'")
                    if len(parts) >= 2:
                        colored_warning = f"  ⚠️  {RED}{BOLD}Signal '{parts[1]}'{RESET}{YELLOW}{warning[len(parts[0]) + len(parts[1]) + 2:]}{RESET}"
                        print(colored_warning)
                    else:
                        print(f"  ⚠️  {YELLOW}{warning}{RESET}")
                else:
                    print(f"  ⚠️  {YELLOW}{warning}{RESET}")
            print(f"{YELLOW}{'='*80}{RESET}\n")

    def _check_unparsed_lines(self, text: str, detected_format: str) -> None:
        """检查可能被跳过的行并发出警告"""
        lines = text.strip().split('\n')
        parsed_names = {s.name.rstrip('0123456789_') if s.name[-1:].isdigit() else s.name
                        for s in self.signals}
        warnings = []

        for line in lines:
            stripped = line.strip()
            if not stripped or stripped.startswith('//') or stripped.startswith('#'):
                continue

            if detected_format == "input":
                # Input Port 格式：检查看起来像 input 声明但没被解析的行
                if re.match(r'^[ \t]*input\b', stripped, re.IGNORECASE):
                    if 'efuse_default_value' not in stripped:
                        warnings.append(f"  Missing 'efuse_default_value' comment: {stripped[:80]}")
                    elif not re.search(r'[，,;]', stripped):
                        warnings.append(f"  Missing comma/semicolon delimiter: {stripped[:80]}")

            elif detected_format == "csv":
                # CSV 格式：检查有逗号但没被解析的行
                parts = stripped.split(',')
                if len(parts) >= 3 and not stripped.lower().startswith('bitwith'):
                    first_col = parts[0].strip()
                    third_col = parts[2].strip()
                    if not ((first_col == '' or re.match(r'^\[\d+:\d+\]$', first_col)) and
                            (third_col.startswith('0x') or third_col.isdigit())):
                        warnings.append(f"  Invalid CSV row: {stripped[:80]}")

            elif detected_format == "hdl":
                # HDL 格式：检查有 logic/input 关键字但没被解析的行
                if re.match(r'^[ \t]*(logic|input|output)\b', stripped, re.IGNORECASE):
                    if not re.search(r'\[\d+:\d+\]', stripped):
                        warnings.append(f"  Missing bit range [msb:lsb]: {stripped[:80]}")

        if warnings:
            RED = '\033[91m'
            YELLOW = '\033[93m'
            BOLD = '\033[1m'
            RESET = '\033[0m'

            print(f"\n{BOLD}{RED}  [WARN] UNPARSED LINE WARNINGS ({len(warnings)}){RESET}")
            print(f"{YELLOW}{'='*80}{RESET}")
            for w in warnings:
                print(f"{YELLOW}{w}{RESET}")
            print(f"{YELLOW}{'='*80}{RESET}\n")

    def parse_signals(self, signal_text: str) -> List[Signal]:
        """
        解析信号列文本，支持多种格式：
        - SystemVerilog: logic [31:0] param_a, param_b;
        - Verilog: input [15:0] data;
        - CSV格式: bitwith,efuse_signals,default_value
        - Input端口格式: input [width] name ， //efuse_default_value: value
        """
        # 清空之前的数据
        self.signals = []
        self.total_width = 0
        self._signal_name_counter = {}  # 重置重复名称计数器

        # 检测格式并解析
        if self._is_csv_format(signal_text):
            self._parse_csv(signal_text)
            self._check_unparsed_lines(signal_text, "csv")
        elif self._is_input_format(signal_text):
            self._parse_input(signal_text)
            self._check_unparsed_lines(signal_text, "input")
        else:
            self._parse_hdl(signal_text)
            self._check_unparsed_lines(signal_text, "hdl")

        # 验证信号值的合法性
        self._validate_signal_values()

        return self.signals

    def _is_input_format(self, text: str) -> bool:
        """检测是否为 input 端口格式"""
        lines = text.strip().split('\n')
        for line in lines[:10]:  # 检查前10行
            line = line.strip()
            if line and not line.startswith('//'):
                # 特征：行以 input 开头，包含分隔符（，,;）和 efuse_default_value 注释
                if line.startswith('input'):
                    if re.search(r'[，,;]', line) and 'efuse_default_value' in line:
                        return True
        return False

    def _parse_input(self, text: str) -> None:
        """解析 input 端口格式信号列表

        格式：input [width] name ，  //efuse_default_value: value
        支持：input [7:0] name ，  //efuse_default_value: 0x123
              input [4-1:0] name ，  //efuse_default_value: 0x8
              input name ，  //efuse_default_value: 0x0
        """
        # 匹配 input 声明行
        # 位宽支持：[msb:lsb] 或 [msb_expr:lsb] 其中 msb_expr 可以是 "N-1" 格式
        # 支持格式：input [7:0] name, 或 input logic [7:0] name, 或 input name,
        pattern = re.compile(
            r"^[ \t]*input\s*"
            r"(?:logic\s+)?"  # 可选的 logic 关键字
            r"(?:\[([\d\s\-]+):(\d+)\]\s*)?"  # 可选的 [msb:lsb] 宽度，msb 支持表达式如 "4-1"
            r"(\w+)\s*"  # 信号名
            r"[，,;]"  # 中文逗号、英文逗号或分号
            r".*?"
            r"efuse_default_value:\s*(0x[0-9a-fA-F]+|\d+)",  # value
            re.MULTILINE | re.IGNORECASE
        )

        for match in pattern.finditer(text):
            msb_str = match.group(1)
            lsb_str = match.group(2)
            name = match.group(3)
            value_str = match.group(4)

            # 计算宽度，判断是否有显式位宽声明
            if msb_str and lsb_str:
                # 解析 msb（可能是表达式如 "4-1" 或简单数字）
                msb_val = self._eval_simple_expr(msb_str.strip())
                lsb_val = int(lsb_str)
                width = msb_val - lsb_val + 1
                explicit_width = True  # 显式声明了位宽（如 [4-1:0] 或 [0:0]）
            else:
                width = 1
                explicit_width = False  # 隐式 1-bit（如 input name,）

            # 解析 value
            value = self._parse_value(value_str) if value_str else 0

            # 生成唯一的信号名（处理重复）
            unique_name = self._get_unique_signal_name(name)

            signal = Signal(
                name=unique_name,
                width=width,
                signal_type=SignalType.LOGIC,
                value=value,
                explicit_width=explicit_width,
                is_reserved=self._is_reserved(name)
            )
            self.signals.append(signal)
            self.total_width += width

    def _eval_simple_expr(self, expr: str) -> int:
        """计算简单算术表达式（如 "4-1", "8-1" 等）"""
        expr = expr.strip()
        # 尝试直接解析为整数
        try:
            return int(expr)
        except ValueError:
            pass
        
        # 处理简单的减法表达式如 "4-1", "10-1"
        if '-' in expr:
            parts = expr.split('-')
            if len(parts) == 2:
                try:
                    left = int(parts[0].strip())
                    right = int(parts[1].strip())
                    return left - right
                except ValueError:
                    pass
        
        # 如果无法解析，报错
        raise ValueError(f"Cannot parse bit width expression: {expr}")

    def _is_csv_format(self, text: str) -> bool:
        """检测是否为CSV格式"""
        lines = text.strip().split('\n')
        for line in lines[:5]:  # 检查前几行
            line = line.strip()
            if line and not line.startswith(';'):
                # CSV特征：逗号分隔，且有3列（bitwidth, name, value）
                parts = line.split(',')
                if len(parts) >= 3:
                    # 检查是否符合CSV格式特征
                    # 第一列可能是空或 [msb:lsb] 格式
                    # 第三列是 0x 开头的十六进制或数字
                    first_col = parts[0].strip()
                    third_col = parts[2].strip()
                    if (first_col == '' or re.match(r'^\[\d+:\d+\]$', first_col)) and \
                       (third_col.startswith('0x') or third_col.isdigit()):
                        return True
        return False

    def _parse_csv(self, text: str) -> None:
        """解析CSV格式信号列表"""
        lines = text.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith(';') or line.lower().startswith('bitwith'):
                continue  # 跳过空行、注释和表头
            
            parts = line.split(',')
            if len(parts) < 3:
                continue
            
            bitwidth_str = parts[0].strip()
            name = parts[1].strip()
            value_str = parts[2].strip()
            
            # 解析位宽
            if bitwidth_str == '':
                # 单bit信号，隐含 [0:0]
                high, low = 0, 0
                width = 1
            else:
                # 解析 [msb:lsb] 格式
                match = re.match(r'^\[(\d+):(\d+)\]$', bitwidth_str)
                if match:
                    high = int(match.group(1))
                    low = int(match.group(2))
                    width = high - low + 1
                else:
                    continue  # 格式错误，跳过
            
            # 解析 value
            value = self._parse_value(value_str)

            # 判断是否有显式位宽声明（CSV中，如果有位宽列则视为显式）
            explicit_width = (bitwidth_str != '')

            # 生成唯一的信号名（处理重复）
            unique_name = self._get_unique_signal_name(name)

            signal = Signal(
                name=unique_name,
                width=width,
                signal_type='reg',
                value=value,
                explicit_width=explicit_width,
                is_reserved=self._is_reserved(name)
            )
            self.signals.append(signal)
            self.total_width += width

    def _parse_hdl(self, text: str) -> None:
        """解析HDL格式信号列表"""
        # 关键模式：匹配行首的 "logic [high:low] signal_name [optional_value]"
        # 避免匹配注释中的内容，要求行首是关键字或 [ 开头
        # value 可以是 0x123 或 123 格式
        pattern = re.compile(
            r"^[ \t]*(?:logic|input|output|wire|reg)?\s*"
            r"\[(\d+):(\d+)\]\s+(\w+)"
            r"(?:\s+(0x[0-9a-fA-F]+|\d+))?"  # 可选的 value
            r"[ \t]*(?:\r?\n|$)",  # 行尾
            re.IGNORECASE | re.MULTILINE
        )

        for match in pattern.finditer(text):
            high = int(match.group(1))
            low = int(match.group(2))
            name = match.group(3)
            width = high - low + 1

            # 解析可选的 value
            value_str = match.group(4)
            value = self._parse_value(value_str) if value_str else 0

            # HDL 格式中，如果显式写了 [0:0] 则视为显式 1-bit
            explicit_width = (width == 1)

            # 生成唯一的信号名（处理重复）
            unique_name = self._get_unique_signal_name(name)

            signal = Signal(
                name=unique_name,
                width=width,
                signal_type=self._detect_type(text),
                value=value,
                explicit_width=explicit_width,
                is_reserved=self._is_reserved(name)
            )
            self.signals.append(signal)
            self.total_width += width

    def _parse_value(self, value_str: str) -> int:
        """解析 value 字符串（支持 0x 十六进制或十进制）"""
        value_str = value_str.strip()
        if value_str.lower().startswith('0x'):
            return int(value_str, 16)
        else:
            return int(value_str)

    def _detect_type(self, line: str) -> SignalType:
        """检测信号类型"""
        if 'logic' in line.lower():
            return SignalType.LOGIC
        elif 'bit' in line.lower():
            return SignalType.BIT
        elif 'wire' in line.lower():
            return SignalType.WIRE
        elif 'reg' in line.lower():
            return SignalType.REG
        return SignalType.LOGIC
    
    def _parse_flexible(self, text: str) -> List[Signal]:  # pragma: no cover
        """灵活解析 - 支持多种格式"""
        signals = []
        
        # 格式: name width
        pattern_simple = re.compile(r"(\w+)\s+(\d+)(?:\s*[,;]|$)")
        
        for match in pattern_simple.finditer(text):
            name = match.group(1)
            width = int(match.group(2))
            signals.append(Signal(name=name, width=width))
        
        return signals
    
    def analyze_segments(self, strategy: str = "auto") -> Dict[str, Dict]:
        """
        分析并生成分段方案
        
        策略:
        - "auto": 自动按信号宽度分段
        - "equal": 均分
        - "functional": 按功能分组
        """
        segments = {}
        current_bit = 0
        
        if strategy == "auto":
            # 按信号顺序生成分段
            for i, signal in enumerate(self.signals):
                segments[signal.name] = {
                    'start_bit': current_bit,
                    'width': signal.width,
                    'end_bit': current_bit + signal.width - 1,
                    'index': i,
                    'signal': signal,
                    'value': signal.value,
                    'is_reserved': signal.is_reserved
                }
                current_bit += signal.width

        elif strategy == "equal":
            # 均分到相等大小
            num_segments = len(self.signals)
            segment_size = self.total_width // num_segments

            for i, signal in enumerate(self.signals):
                segments[signal.name] = {
                    'start_bit': i * segment_size,
                    'width': segment_size,
                    'end_bit': i * segment_size + segment_size - 1,
                    'index': i,
                    'signal': signal,
                    'value': signal.value,
                    'is_reserved': signal.is_reserved
                }
        
        return segments

    def _generate_rhs(self, start: int, end: int, width: int,
                      value: int, input_signal: str) -> str:
        """
        根据 value 生成 RHS 表达式（参考 Perl 脚本的 efuse 取反逻辑）

        逻辑:
        - value == 0: 直接取原值
        - value != 0:
          - width == 1: 整体取反 ~
          - width > 1: 逐 bit 取反，根据 value 的每一位决定是否取反该 bit
            - value bit == 0: 取原值
            - value bit == 1: 取反 ~
        """
        if value == 0:
            # 原值，不取反
            if start == end:
                return f"{input_signal}[{start}]"
            else:
                return f"{input_signal}[{end}:{start}]"
        elif width == 1:
            # 整体取反
            return f"~{input_signal}[{start}]"
        else:
            # 逐 bit 取反
            bits = []
            # 从高位到低位
            for bit_pos in range(width - 1, -1, -1):
                bit_value = (value >> bit_pos) & 1
                bit_input = start + bit_pos
                if bit_value == 0:
                    bits.append(f"{input_signal}[{bit_input}]")
                else:
                    bits.append(f"~{input_signal}[{bit_input}]")
            return "{" + ", ".join(bits) + "}"

    def generate_sv_module(self, segments: Dict,
                          module_name: str = "param_distributor",
                          input_signal: str = "chip_param") -> str:
        """生成组合逻辑分发器 module"""

        code = ""

        # 计算总位宽（包含 reserved 信号的位宽）
        max_end = max(seg['end_bit'] for seg in segments.values())
        total_width = max_end + 1

        code += f"module {module_name} (\n"
        code += f"    input  logic [{total_width-1}:0]  {input_signal},\n"
        code += "\n    // Output segments\n"

        # 生成输出端口（跳过 reserved 信号）
        sorted_segs = sorted(segments.items(),
                            key=lambda x: x[1]['start_bit'])

        # 过滤掉 reserved 信号
        non_reserved = [(name, seg) for name, seg in sorted_segs if not seg.get('is_reserved', False)]

        for i, (name, seg) in enumerate(non_reserved):
            width = seg['width']
            signal_obj = seg.get('signal')

            # 根据是否显式声明位宽决定格式
            if width == 1 and signal_obj and not signal_obj.explicit_width:
                # 隐式 1-bit: 不显示 [0:0]
                code += f"    output logic                   {name}"
            else:
                # 多位信号 或 显式 1-bit
                bit_range = f"[{width-1}:0]"
                code += f"    output logic {bit_range}      {name}"

            code += "\n" if i == len(non_reserved) - 1 else ",\n"

        code += ");\n\n"

        # 生成组合逻辑（跳过 reserved 信号）
        code += "    // Combinational logic distribution\n"
        code += "    always_comb begin\n"

        for name, seg in non_reserved:
            start = seg['start_bit']
            width = seg['width']
            value = seg.get('value', 0)

            rhs = self._generate_rhs(start, start + width - 1, width, value, input_signal)
            # 格式化 value 注释
            if value == 0:
                value_comment = "0x0"
            elif value < 16:
                value_comment = f"0x{value:x}"
            else:
                value_comment = f"0x{value:X}"
            code += f"        {name} = {rhs}; // value: {value_comment}\n"

        code += "    end\n\n"
        code += "endmodule\n"

        return code
    
    def generate_struct_definition(self, segments: Dict,
                                   struct_name: str = "param_struct_t") -> str:
        """生成 packed struct 定义"""

        code = f"typedef struct packed {{\n"

        sorted_segs = sorted(segments.items(),
                            key=lambda x: x[1]['start_bit'])

        for name, seg in sorted_segs:
            if seg.get('is_reserved', False):
                continue  # 跳过 reserved 信号
            width = seg['width']
            if width == 1:
                code += f"    logic              {name};\n"
            else:
                code += f"    logic [{width-1}:0]  {name};\n"

        code += f"}} {struct_name};\n"

        return code
    
    def generate_assignment_logic(self, segments: Dict,
                                  input_signal: str = "chip_param") -> str:
        """生成分配逻辑（组合）"""

        code = "// Combinational assignment logic\n"
        code += "always_comb begin\n"

        sorted_segs = sorted(segments.items(),
                            key=lambda x: x[1]['start_bit'])

        for name, seg in sorted_segs:
            if seg.get('is_reserved', False):
                continue  # 跳过 reserved 信号
            start = seg['start_bit']
            width = seg['width']
            value = seg.get('value', 0)

            rhs = self._generate_rhs(start, start + width - 1, width, value, input_signal)
            code += f"    {name} = {rhs};\n"

        code += "end\n"

        return code
    
    def print_analysis(self, segments: Dict) -> None:
        """打印分析结果"""
        print("\n" + "="*80)
        print("SIGNAL ANALYSIS & SEGMENTATION")
        print("="*80)
        print(f"\nTotal signals parsed: {len(self.signals)}")
        print(f"Total parameter width: {self.total_width} bits\n")
        
        print("Segment Map:")
        print("-" * 95)
        print(f"{'Signal Name'.ljust(25)} {'Bit Range'.rjust(15)}  {'Width'.rjust(4)}b  {'Value'.rjust(8)}  {'Invert Type'.rjust(12)}")
        print("-" * 95)

        sorted_segs = sorted(segments.items(),
                            key=lambda x: x[1]['start_bit'])

        for name, seg in sorted_segs:
            start = seg['start_bit']
            end = seg['end_bit']
            width = seg['width']
            value = seg.get('value', 0)
            is_reserved = seg.get('is_reserved', False)

            # 确定取反类型
            if is_reserved:
                invert_type = "reserved"
            elif value == 0:
                invert_type = "none"
            elif width == 1:
                invert_type = "full (~)"
            else:
                invert_type = "per-bit"

            value_hex = f"0x{value:X}" if value else "0"

            if start == end:
                bit_range = f"[{start}]"
            else:
                bit_range = f"[{end}:{start}]"

            display_name = f"{name} (R)" if is_reserved else name
            print(f"{display_name.ljust(25)} {bit_range.rjust(15)}  {width:4d}  {value_hex.rjust(8)}  {invert_type.rjust(12)}")

        print("-" * 95)
        print("="*80 + "\n")

    def generate_byte_table(self, segments: Dict) -> str:
        """
        生成 Byte/Bit 映射表格（参考 eFuse/寄存器手册格式）
        
        格式示例:
        Byte 0    7    6    5    4    3    2    1    0
        Name      LOT_ID_DIGIT_1[1:0]       LOT_ID_DIGIT_0[5:0]
        """
        if not segments:
            return ""
        
        # 按 start_bit 排序信号
        sorted_sigs = sorted(segments.items(), key=lambda x: x[1]['start_bit'])
        
        # 计算总字节数
        max_bit = max(seg['end_bit'] for seg in segments.values())
        total_bytes = (max_bit // 8) + 1
        
        lines = []
        lines.append("\n" + "="*80)
        lines.append("BYTE/BIT MAPPING TABLE")
        lines.append("="*80 + "\n")
        
        # 列宽设置 - 8字符足够显示大部分信号名如 "LOT_ID_0[5:0]"
        COL_WIDTH = 8
        
        for byte_idx in range(total_bytes):
            byte_start = byte_idx * 8
            byte_end = byte_start + 7
            
            # 标题行
            header = f"Byte {byte_idx}".ljust(8)
            for bit in range(7, -1, -1):
                header += str(bit).center(COL_WIDTH)
            lines.append(header)
            
            # 构建 Name 行 - 每个信号在它所占的第一个列位置显示
            cells = [None] * 8
            
            for name, seg in sorted_sigs:
                seg_start = seg['start_bit']
                seg_end = seg['end_bit']
                
                if seg_end < byte_start or seg_start > byte_end:
                    continue
                
                cover_start = max(seg_start, byte_start) - byte_start
                cover_end = min(seg_end, byte_end) - byte_start
                
                sig_bit_start = max(seg_start, byte_start) - seg_start
                sig_bit_end = min(seg_end, byte_end) - seg_start
                
                col_start = 7 - cover_end
                
                if seg_start == seg_end:
                    text = name
                else:
                    if sig_bit_start == sig_bit_end:
                        text = f"{name}[{sig_bit_end}]"
                    else:
                        text = f"{name}[{sig_bit_end}:{sig_bit_start}]"
                
                # 只在第一个空闲位置显示
                if cells[col_start] is None:
                    cells[col_start] = text
            
            # 生成 Name 行
            name_line = "Name".ljust(8)
            for col in range(8):
                content = cells[col] if cells[col] else ""
                name_line += content.ljust(COL_WIDTH)
            lines.append(name_line)
            lines.append("")
            
        lines.append("="*80 + "\n")
        return "\n".join(lines)

    def print_byte_table(self, segments: Dict) -> None:
        """打印 Byte/Bit 映射表格"""
        print(self.generate_byte_table(segments))

    def generate_byte_table_excel(self, segments: Dict, output_file: str) -> None:
        """
        生成 Byte/Bit 映射表格的 Excel 文件（支持多比特合并单元格）
        
        Args:
            segments: 信号分段字典
            output_file: 输出文件路径 (.xlsx)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("Warning: openpyxl not available, skipping Excel export")
            return
        
        if not segments:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Byte_Bit_Mapping"
        
        # 设置样式
        header_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=14)
        cell_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # 按 start_bit 排序信号
        sorted_sigs = sorted(segments.items(), key=lambda x: x[1]['start_bit'])
        
        # 定义颜色池（柔和的背景色）
        color_palette = [
            'FFE6E6',  # 浅红
            'E6F7FF',  # 浅蓝
            'E6FFE6',  # 浅绿
            'FFF5E6',  # 浅橙
            'F5E6FF',  # 浅紫
            'E6FFFF',  # 浅青
            'FFFFCC',  # 浅黄
            'FFE6F5',  # 浅粉
            'E6FFF5',  # 浅薄荷
            'FFF0E6',  # 浅桃
            'E8E6FF',  # 浅薰衣草
            'F0FFE6',  # 浅柠檬绿
        ]
        
        # 为每个信号名称分配颜色
        signal_names = list(dict.fromkeys(name for name, _ in sorted_sigs))  # 按首次出现顺序去重
        signal_color_map = {}
        for i, name in enumerate(signal_names):
            signal_color_map[name] = color_palette[i % len(color_palette)]
        
        # 缓存 fill 对象避免重复创建
        fill_cache = {}

        # 计算总字节数
        max_bit = max(seg['end_bit'] for seg in segments.values())
        total_bytes = (max_bit // 8) + 1
        
        # 写入标题行
        ws['A1'] = "Byte/Bit Mapping Table"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        
        current_row = 3
        
        for byte_idx in range(total_bytes):
            byte_start = byte_idx * 8
            byte_end = byte_start + 7
            
            # Byte 标题行
            ws.cell(row=current_row, column=1, value=f"Byte {byte_idx}")
            ws.cell(row=current_row, column=1).font = header_font
            
            for bit, col in zip(range(7, -1, -1), range(2, 10)):
                cell = ws.cell(row=current_row, column=col, value=bit)
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            current_row += 1
            
            # Name 行
            ws.cell(row=current_row, column=1, value="Name")
            ws.cell(row=current_row, column=1).font = header_font

            # 处理每个 bit 列（2-9 对应 bit 7-0）
            col_idx = 2  # 从第 2 列开始（bit 7）

            while col_idx <= 9:
                # 计算当前列对应的 bit 位置
                bit_pos = 9 - col_idx  # col 2 -> bit 7, col 3 -> bit 6, ..., col 9 -> bit 0
                bit_in_byte = 7 - bit_pos

                # 查找该 bit 位置所属的信号
                found_signal = None
                found_seg = None
                sig_cover_start = 0
                sig_cover_end = 0
                sig_bit_start = 0
                sig_bit_end = 0

                for name, seg in sorted_sigs:
                    seg_start = seg['start_bit']
                    seg_end = seg['end_bit']

                    if seg_end < byte_start or seg_start > byte_end:
                        continue

                    # 计算信号在该字节内的覆盖范围
                    cover_start = max(seg_start, byte_start) - byte_start
                    cover_end = min(seg_end, byte_end) - byte_start

                    # 检查当前 bit 是否属于这个信号
                    if cover_start <= bit_in_byte <= cover_end:
                        found_signal = name
                        found_seg = seg
                        sig_cover_start = cover_start
                        sig_cover_end = cover_end
                        sig_bit_start = max(seg_start, byte_start) - seg_start
                        sig_bit_end = min(seg_end, byte_end) - seg_start
                        break

                if found_signal:
                    # 计算需要合并的列数
                    cover_bits = sig_cover_end - sig_cover_start + 1
                    start_col = col_idx
                    end_col = col_idx + cover_bits - 1

                    # 生成信号文本
                    if seg_start == seg_end:
                        text = found_signal
                    else:
                        if sig_bit_start == sig_bit_end:
                            text = f"{found_signal}[{sig_bit_end}]"
                        else:
                            text = f"{found_signal}[{sig_bit_end}:{sig_bit_start}]"

                    # 获取或创建颜色 fill 对象（reserved 信号使用灰色）
                    if found_seg.get('is_reserved', False):
                        signal_color = RESERVED_COLOR
                    else:
                        signal_color = signal_color_map.get(found_signal)
                    if signal_color:
                        if signal_color not in fill_cache:
                            fill_cache[signal_color] = PatternFill(start_color=signal_color, end_color=signal_color, fill_type='solid')
                        color_fill = fill_cache[signal_color]
                    else:
                        color_fill = None

                    # 写入第一个单元格并合并
                    ws.cell(row=current_row, column=start_col, value=text)
                    ws.cell(row=current_row, column=start_col).font = cell_font
                    ws.cell(row=current_row, column=start_col).border = border
                    ws.cell(row=current_row, column=start_col).alignment = left_align
                    if color_fill:
                        ws.cell(row=current_row, column=start_col).fill = color_fill

                    if cover_bits > 1:
                        # 合并单元格
                        merge_range = f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{current_row}"
                        ws.merge_cells(merge_range)

                    # 跳到下一个信号之后的位置
                    col_idx = end_col + 1
                else:
                    # 空单元格
                    ws.cell(row=current_row, column=col_idx, value="")
                    ws.cell(row=current_row, column=col_idx).border = border
                    col_idx += 1

            current_row += 1

            # Default Value 行 - 显示 efuse_default_value
            ws.cell(row=current_row, column=1, value="Default")
            ws.cell(row=current_row, column=1).font = header_font

            # 处理每个 bit 列的默认值
            col_idx = 2

            while col_idx <= 9:
                bit_pos = 9 - col_idx
                bit_in_byte = 7 - bit_pos

                found_signal = None
                found_seg = None
                sig_cover_start = 0
                sig_cover_end = 0

                for name, seg in sorted_sigs:
                    seg_start = seg['start_bit']
                    seg_end = seg['end_bit']

                    if seg_end < byte_start or seg_start > byte_end:
                        continue

                    cover_start = max(seg_start, byte_start) - byte_start
                    cover_end = min(seg_end, byte_end) - byte_start

                    if cover_start <= bit_in_byte <= cover_end:
                        found_signal = name
                        found_seg = seg
                        sig_cover_start = cover_start
                        sig_cover_end = cover_end
                        break

                if found_signal and found_seg:
                    cover_bits = sig_cover_end - sig_cover_start + 1
                    start_col = col_idx
                    end_col = col_idx + cover_bits - 1

                    # 生成 value 文本
                    value = found_seg.get('value', 0)
                    if value == 0:
                        value_text = "0x0"
                    elif value < 16:
                        value_text = f"0x{value:x}"
                    else:
                        value_text = f"0x{value:X}"

                    # 获取颜色（reserved 信号使用灰色）
                    if found_seg.get('is_reserved', False):
                        signal_color = RESERVED_COLOR
                    else:
                        signal_color = signal_color_map.get(found_signal)
                    if signal_color:
                        if signal_color not in fill_cache:
                            fill_cache[signal_color] = PatternFill(start_color=signal_color, end_color=signal_color, fill_type='solid')
                        color_fill = fill_cache[signal_color]
                    else:
                        color_fill = None

                    # 写入 value
                    ws.cell(row=current_row, column=start_col, value=value_text)
                    ws.cell(row=current_row, column=start_col).font = cell_font
                    ws.cell(row=current_row, column=start_col).border = border
                    ws.cell(row=current_row, column=start_col).alignment = center_align
                    if color_fill:
                        ws.cell(row=current_row, column=start_col).fill = color_fill

                    if cover_bits > 1:
                        merge_range = f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{current_row}"
                        ws.merge_cells(merge_range)

                    col_idx = end_col + 1
                else:
                    ws.cell(row=current_row, column=col_idx, value="")
                    ws.cell(row=current_row, column=col_idx).border = border
                    col_idx += 1

            current_row += 1
            current_row += 1  # 空行
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 15
        
        # 保存
        wb.save(output_file)
        print(f"✓ Excel table generated: {output_file}")


class DistributorGenerator:
    """分发器代码生成器"""
    
    def __init__(self, parser: DynamicSignalParser):
        self.parser = parser
    
    def generate_complete_sv(self, output_file: str = "param_distributor.sv") -> str:
        """生成完整的 SV 文件"""
        
        segments = self.parser.analyze_segments(strategy="auto")
        
        code = ""
        
        # Struct 定义
        code += self.parser.generate_struct_definition(segments) + "\n\n"
        
        # Module 定义
        code += self.parser.generate_sv_module(segments) + "\n"
        
        return code
    
    def generate_with_options(self, segments: Dict,
                             include_struct: bool = True,
                             include_comments: bool = True,
                             module_name: str = "param_distributor") -> str:
        """按选项生成代码"""

        code = ""

        if include_comments:
            pass

        if include_struct:
            code += self.parser.generate_struct_definition(segments) + "\n\n"

        code += self.parser.generate_sv_module(segments, module_name=module_name) + "\n"

        return code


def main():  # pragma: no cover
    """主函数 - 演示用法"""

    import argparse

    parser = argparse.ArgumentParser(
        description="Dynamic Signal Parser & Parameter Distributor (with Value-based Inversion)"
    )
    parser.add_argument('-i', '--input', type=str, required=True,
                       help='Input signal list file or text')
    parser.add_argument('-o', '--output', type=str, default='param_distributor.sv',
                       help='Output SV file (basename only, will be placed in output_dir)')
    parser.add_argument('-d', '--output-dir', type=str, default='generated',
                       help='Output directory for generated files (default: generated)')
    parser.add_argument('-s', '--strategy', choices=['auto', 'equal', 'functional'],
                       default='auto', help='Segmentation strategy')
    parser.add_argument('--struct', action='store_true', default=True,
                       help='Include struct definition')
    parser.add_argument('-q', '--quiet', action='store_true', default=False,
                       help='Suppress verbose console output (Segment Map, Byte Table)')

    args = parser.parse_args()
    
    # 确保输出目录存在
    os.makedirs(args.output_dir, exist_ok=True)

    # 读取输入信号
    try:
        with open(args.input, 'r', encoding='utf-8') as f:
            signal_text = f.read()
    except FileNotFoundError:
        signal_text = args.input
    except UnicodeDecodeError:
        # 尝试其他编码
        with open(args.input, 'r', encoding='gbk') as f:
            signal_text = f.read()

    # 解析信号
    dsp = DynamicSignalParser()
    signals = dsp.parse_signals(signal_text)

    if not signals:
        print("❌ No signals parsed. Check input format.")
        print("\nExpected formats:")
        print("  - logic [31:0] param_a          (value = 0, no inversion)")
        print("  - logic [31:0] param_b 0x123    (value != 0, per-bit inversion)")
        print("  - logic [0:0]  valid 0x1       (single bit, full inversion)")
        print("\nInversion Rules:")
        print("  - value == 0    : Direct assignment (no inversion)")
        print("  - width == 1    : Full inversion (~)")
        print("  - width > 1     : Per-bit inversion based on value bits")
        return

    # 分析分段
    segments = dsp.analyze_segments(strategy=args.strategy)
    if not args.quiet:
        dsp.print_analysis(segments)
        dsp.print_byte_table(segments)
    
    # 生成 Excel Byte/Bit 表格（使用日期时间避免文件占用）
    output_sv_path = os.path.join(args.output_dir, args.output)
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_output = os.path.join(args.output_dir, args.output.replace('.sv', f'_byte_table_{timestamp_str}.xlsx'))
    dsp.generate_byte_table_excel(segments, excel_output)

    # 生成代码
    gen = DistributorGenerator(dsp)
    # 从输出文件名提取 module_name（去掉 .sv 后缀）
    module_name = os.path.splitext(args.output)[0]
    sv_code = gen.generate_with_options(
        segments,
        include_struct=args.struct,
        module_name=module_name
    )

    # 写入文件
    with open(output_sv_path, 'w') as f:
        f.write(sv_code)

    print(f"✓ Generated: {output_sv_path}\n")

    # 显示预览
    print("Preview:")
    print("-" * 80)
    print(sv_code[:1000])
    if len(sv_code) > 1000:
        print("\n... (truncated)")


if __name__ == "__main__":  # pragma: no cover
    main()
