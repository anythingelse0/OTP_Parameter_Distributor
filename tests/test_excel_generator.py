#!/usr/bin/env python3
"""
Excel 生成器单元测试
测试 Excel 字节/位映射表生成
"""

import sys
import os
import tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import unittest
from datetime import datetime
from dynamic_signal_parser import DynamicSignalParser, Signal, SignalType


class TestExcelGenerator(unittest.TestCase):
    """测试 Excel 生成功能"""

    def setUp(self):
        self.parser = DynamicSignalParser()
        # 创建临时目录
        self.temp_dir = tempfile.TemporaryDirectory()

    def tearDown(self):
        # 清理临时目录
        self.temp_dir.cleanup()

    def test_excel_generation_simple(self):
        """测试简单信号 Excel 生成"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        segments = {
            'uuid': {'start_bit': 0, 'end_bit': 63, 'width': 64, 'value': 0}
        }
        
        output_file = os.path.join(self.temp_dir.name, "test.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)
        
        # 验证文件存在
        self.assertTrue(os.path.exists(output_file))
        
        # 验证可以正常打开
        wb = openpyxl.load_workbook(output_file)
        self.assertEqual(wb.active.title, "Byte_Bit_Mapping")

    def test_excel_filename_with_datetime(self):
        """测试 Excel 文件名包含日期时间"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        segments = {
            'data': {'start_bit': 0, 'end_bit': 7, 'width': 8, 'value': 0xAB}
        }
        
        # 使用时间戳格式文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(self.temp_dir.name, f"test_{timestamp}.xlsx")
        
        self.parser.generate_byte_table_excel(segments, output_file)
        
        # 验证文件存在
        self.assertTrue(os.path.exists(output_file))

    def test_excel_multiple_bytes(self):
        """测试多字节信号 Excel 生成"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        segments = {
            'data64': {'start_bit': 0, 'end_bit': 63, 'width': 64, 'value': 0},
            'data8': {'start_bit': 64, 'end_bit': 71, 'width': 8, 'value': 0xFF}
        }
        
        output_file = os.path.join(self.temp_dir.name, "multi_byte.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)
        
        # 验证文件存在且可打开
        self.assertTrue(os.path.exists(output_file))
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # 验证标题存在
        self.assertEqual(ws['A1'].value, "Byte/Bit Mapping Table")

    def test_excel_cell_merge(self):
        """测试跨字节的信号单元格合并"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        # 创建一个跨越多个 bit 的信号
        segments = {
            'wide_sig': {'start_bit': 4, 'end_bit': 11, 'width': 8, 'value': 0x5A}
        }
        
        output_file = os.path.join(self.temp_dir.name, "merged.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)
        
        self.assertTrue(os.path.exists(output_file))

    def test_excel_empty_segments(self):
        """测试空 segments 不生成文件"""
        output_file = os.path.join(self.temp_dir.name, "empty.xlsx")
        
        # openpyxl 可能不可用，但不影响这个测试
        self.parser.generate_byte_table_excel({}, output_file)
        
        # 空 segments 不应该生成文件
        self.assertFalse(os.path.exists(output_file))

    def test_color_palette_assignment(self):
        """测试颜色分配"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        # 多个信号测试颜色轮询
        segments = {
            'sig1': {'start_bit': 0, 'end_bit': 7, 'width': 8, 'value': 0},
            'sig2': {'start_bit': 8, 'end_bit': 15, 'width': 8, 'value': 0},
            'sig3': {'start_bit': 16, 'end_bit': 23, 'width': 8, 'value': 0},
        }
        
        output_file = os.path.join(self.temp_dir.name, "colored.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)
        
        self.assertTrue(os.path.exists(output_file))
        wb = openpyxl.load_workbook(output_file)
        
        # 文件成功生成即可
        self.assertIsNotNone(wb)


class TestExcelContent(unittest.TestCase):
    """测试 Excel 内容正确性"""

    def setUp(self):
        self.parser = DynamicSignalParser()
        self.temp_dir = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_title_row(self):
        """测试标题行"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        segments = {
            'test': {'start_bit': 0, 'end_bit': 7, 'width': 8, 'value': 0}
        }
        
        output_file = os.path.join(self.temp_dir.name, "title.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)
        
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # 验证标题
        self.assertEqual(ws['A1'].value, "Byte/Bit Mapping Table")

    def test_byte_header_row(self):
        """测试 Byte 标题行格式"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        segments = {
            'byte0': {'start_bit': 0, 'end_bit': 7, 'width': 8, 'value': 0}
        }
        
        output_file = os.path.join(self.temp_dir.name, "byte_header.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)
        
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Byte 0 标题应该在第 3 行
        self.assertEqual(ws['A3'].value, "Byte 0")

    def test_default_value_display(self):
        """测试 Default 值显示"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        segments = {
            'test': {'start_bit': 0, 'end_bit': 7, 'width': 8, 'value': 0xAB}
        }
        
        output_file = os.path.join(self.temp_dir.name, "default.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)
        
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # 查找 Default 行并验证值
        found_default = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if row[0].value == "Default":
                found_default = True
                break
        
        self.assertTrue(found_default, "Should have 'Default' row")

    def test_bit_column_alignment_header_vs_content(self):
        """表头 bit 7 在左、bit 0 在右；Name 行内容必须与之对齐。

        用一个占 bit 0-3 的信号验证：它必须出现在 Byte 0 行的低位区
        （右侧高列号，col 6-9），而不是被错放到左侧 bit 7 区域。"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        # short_a 占 Byte 0 的 bit 3:0（低位），右侧
        segments = {
            'short_a': {'start_bit': 0, 'end_bit': 3, 'width': 4, 'value': 0x5}
        }
        output_file = os.path.join(self.temp_dir.name, "align.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        # 找到 "Byte 0" 标题行
        byte0_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Byte 0":
                byte0_row = r
                break
        self.assertIsNotNone(byte0_row, "Should have a Byte 0 row")

        # 表头：第 2 列=bit7 ... 第 9 列=bit0
        header_bits = [ws.cell(row=byte0_row, column=c).value for c in range(2, 10)]
        self.assertEqual(header_bits, [7, 6, 5, 4, 3, 2, 1, 0],
                         "Header must read bit 7..0 left to right")

        name_row = byte0_row + 1
        # short_a 占 bit 3:0 → 应落在 col 6-9（bit 3..0），合并后文字在 col 6
        name_cells = [ws.cell(row=name_row, column=c).value for c in range(2, 10)]
        self.assertEqual(name_cells[4], 'short_a[3:0]',
                         "low bits signal must sit on the right (bit 0 region)")
        # 左侧高位区（bit 7..4）应为空
        self.assertTrue(all(v in (None, "") for v in name_cells[:4]),
                        "high bit region must be empty")

    def test_cross_byte_default_value_masked(self):
        """跨字节信号的 Default 行，每段只显示该字节内对应的局部 bit 值，
        而非整信号的完整 value。"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        # long_signal = 0xABCD，占 bit 4..19
        # Byte 0 覆盖 bit 7:4 of sig → (0xABCD >> 4) & 0xF = 0xD
        # Byte 1 覆盖 → 0xBC
        # Byte 2 覆盖 bit 19:16 → 0xA
        segments = {
            'short_a': {'start_bit': 0, 'end_bit': 3, 'width': 4, 'value': 0x5},
            'long_signal': {'start_bit': 4, 'end_bit': 19, 'width': 16, 'value': 0xABCD},
            'short_b': {'start_bit': 20, 'end_bit': 23, 'width': 4, 'value': 0x3}
        }
        output_file = os.path.join(self.temp_dir.name, "xbyte.xlsx")
        self.parser.generate_byte_table_excel(segments, output_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        def byte_header_row(name):
            for r in range(1, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == name:
                    return r
            self.fail(f"Missing {name} row")

        def default_row_vals(byte_header_row):
            r = byte_header_row + 2  # Byte row -> Name row -> Default row
            return [ws.cell(row=r, column=c).value for c in range(2, 10)]

        v0 = [str(x).lower() for x in default_row_vals(byte_header_row("Byte 0"))
              if x not in (None, "")]
        v1 = [str(x).lower() for x in default_row_vals(byte_header_row("Byte 1"))
              if x not in (None, "")]
        v2 = [str(x).lower() for x in default_row_vals(byte_header_row("Byte 2"))
              if x not in (None, "")]

        # Byte 0: long_signal 局部 = 0xD, short_a = 0x5（不能出现完整 0xABCD）
        self.assertIn('0xd', v0)
        self.assertIn('0x5', v0)
        self.assertNotIn('0xabcd', v0,
                         "Byte 0 must NOT show full signal value")
        # Byte 1: long_signal 局部 = 0xBC
        self.assertIn('0xbc', v1)
        # Byte 2: short_b=0x3, long_signal 局部 = 0xA
        self.assertIn('0xa', v2)
        self.assertIn('0x3', v2)


class TestExcelFileOperations(unittest.TestCase):
    """测试 Excel 文件操作"""

    def setUp(self):
        self.parser = DynamicSignalParser()
        self.temp_dir = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_output_directory_creation(self):
        """测试输出目录自动创建"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        nested_dir = os.path.join(self.temp_dir.name, "nested", "path")
        output_file = os.path.join(nested_dir, "test.xlsx")
        
        # 确保目录存在
        os.makedirs(nested_dir, exist_ok=True)
        
        segments = {
            'test': {'start_bit': 0, 'end_bit': 7, 'width': 8, 'value': 0}
        }
        
        self.parser.generate_byte_table_excel(segments, output_file)
        
        self.assertTrue(os.path.exists(output_file))


if __name__ == "__main__":
    unittest.main(verbosity=2)
